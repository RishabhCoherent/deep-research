"""
Extract Market Estimate (ME) data from an XLSX file.

Discovery-based parser — no hardcoded row numbers, column positions,
segment names, region lists, or year ranges. All structure is inferred
from the spreadsheet content itself.

Universal patterns relied upon (structural, not content-specific):
  - Header row: the row where most cell values are 4-digit year numbers
  - Three column groups: Forecast years, YoY Growth years, Percentage Share years
  - CAGR column: header cell contains "CAGR" (period parsed from text)
  - Three row sections: Market Volume, Pricing, Market Value
  - Segment summary rows: labels starting with "By " (e.g., "By Product Type")
  - Geographic summary rows: labels like "Region", "Country", etc.
  - Data sheets vs template sheets: data sheets have "Market Volume"/"Market Value" rows
  - Global sheet: the data sheet whose geographic entries match other sheet names
"""

import re
import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


# ─── Header / Column Discovery ──────────────────────────────────────────────


def _is_year(val) -> bool:
    """Check if a value looks like a calendar year."""
    if val is None:
        return False
    try:
        n = int(float(val))
        return 1990 <= n <= 2099
    except (ValueError, TypeError):
        return False


def _discover_header_row(ws, max_scan: int = 10) -> int | None:
    """Find the header row by scanning for the row with the most year-like values."""
    best_row, best_count = None, 0
    for row in range(1, min(ws.max_row + 1, max_scan + 1)):
        count = sum(1 for col in range(1, ws.max_column + 1) if _is_year(ws.cell(row=row, column=col).value))
        if count > best_count:
            best_row, best_count = row, count
    return best_row if best_count >= 3 else None


def _discover_column_groups(ws, header_row: int) -> dict:
    """Discover all three column groups from the header row.

    Layout: [Forecast years] [CAGR] [gap] [YoY Growth years] [gap] [Percentage Share years]

    Returns:
        {
            "forecast_cols": [(col_idx, year_int), ...],
            "cagr_col": col_idx | None,
            "cagr_label": "CAGR(2025-32)" | None,
            "yoy_cols": [(col_idx, year_int), ...],
            "pct_cols": [(col_idx, year_int), ...],
        }
    """
    cagr_col = None
    cagr_label = None
    year_groups = []
    current_group = []

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value

        if val is not None:
            s = str(val).strip()

            if re.search(r'CAGR', s, re.IGNORECASE):
                if current_group:
                    year_groups.append(current_group)
                    current_group = []
                cagr_col = col
                cagr_label = s
                continue

            if _is_year(val):
                current_group.append((col, int(float(val))))
                continue

        # Non-year, non-CAGR cell (or None) — ends current group
        if current_group:
            year_groups.append(current_group)
            current_group = []

    if current_group:
        year_groups.append(current_group)

    return {
        "forecast_cols": year_groups[0] if len(year_groups) > 0 else [],
        "cagr_col": cagr_col,
        "cagr_label": cagr_label,
        "yoy_cols": year_groups[1] if len(year_groups) > 1 else [],
        "pct_cols": year_groups[2] if len(year_groups) > 2 else [],
    }


def _parse_cagr_key(cagr_label: str | None) -> str:
    """Turn 'CAGR(2025-32)' into 'cagr_2025_2032'. Handles many formats."""
    if not cagr_label:
        return "cagr"
    m = re.search(r'(\d{4})\s*[-–]\s*(\d{2,4})', cagr_label)
    if m:
        start = m.group(1)
        end = m.group(2)
        if len(end) == 2:
            end = start[:2] + end
        return f"cagr_{start}_{end}"
    return "cagr"


# ─── Sheet Classification ───────────────────────────────────────────────────


def _is_data_sheet(ws) -> bool:
    """A data sheet has 'Market Volume' or 'Market Value' somewhere in column A."""
    for row in range(1, min(ws.max_row + 1, 10)):
        val = ws.cell(row=row, column=1).value
        if val and re.search(r'market\s+(volume|value)', str(val), re.IGNORECASE):
            return True
    return False


def _identify_global_sheet(wb, data_sheet_names: list[str]) -> str | None:
    """The global sheet is the one whose geographic entries match other data sheet names."""
    for candidate in data_sheet_names:
        ws = wb[candidate]
        labels = set()
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val:
                labels.add(str(val).strip())

        other_names = {n.strip() for n in data_sheet_names if n != candidate}
        matches = labels & other_names
        if len(matches) >= 2:
            return candidate

    for name in data_sheet_names:
        if re.search(r'global|world|total|summary', name, re.IGNORECASE):
            return name

    return data_sheet_names[0] if data_sheet_names else None


# ─── Row Discovery ──────────────────────────────────────────────────────────


def _discover_row_structure(ws) -> dict:
    """Scan column A and classify every row by its structural role.

    Returns:
        {
            "total_rows": {"volume": row, "value": row},
            "sections": {
                "market_volume": {"start": row, "end": row, "unit": "..."},
                "pricing": {"start": row, "end": row, "unit": "..."},
                "market_value": {"start": row, "end": row, "unit": "..."},
            },
            "summary_rows": {row: "By Product Type", ...},
            "geo_summary_rows": {row: "Region"|"Country"|...},
            "data_rows": [row, ...],
        }
    """
    total_rows = {}
    sections = {}
    summary_rows = {}
    geo_summary_rows = {}
    data_rows = []

    seen_market_volume = False
    seen_market_value = False

    for row in range(1, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if raw is None:
            continue
        label = str(raw).strip()
        lower = label.lower()

        # Total Market Volume / Market Value (first occurrence, exact match)
        if re.match(r'^market\s+volume$', lower) and not seen_market_volume:
            total_rows["volume"] = row
            seen_market_volume = True
            continue
        if re.match(r'^market\s+value$', lower) and not seen_market_value:
            total_rows["value"] = row
            seen_market_value = True
            continue

        # Section headers with units in parentheses
        if re.match(r'^market\s+volume\s*\(', lower):
            unit_match = re.search(r'\((.+?)\)', label)
            sections["market_volume"] = {
                "start": row,
                "unit": unit_match.group(1) if unit_match else "",
            }
            continue
        if re.match(r'^pricing\s*\(', lower):
            unit_match = re.search(r'\((.+?)\)', label)
            sections["pricing"] = {
                "start": row,
                "unit": unit_match.group(1) if unit_match else "",
            }
            continue
        if re.match(r'^market\s+value\s*\(', lower):
            unit_match = re.search(r'\((.+?)\)', label)
            sections["market_value"] = {
                "start": row,
                "unit": unit_match.group(1) if unit_match else "",
            }
            continue

        # Geographic summary headers (Region, Country, etc.)
        if re.match(r'^(region|country|state|province|territory)$', lower):
            geo_summary_rows[row] = label
            continue

        # Segment summary rows: labels starting with "By "
        if re.match(r'^by\s+', label, re.IGNORECASE):
            summary_rows[row] = label
            continue

        # Everything else is a data row
        data_rows.append(row)

    # Compute section end boundaries
    ordered = sorted(sections.items(), key=lambda x: x[1]["start"])
    for i, (_, info) in enumerate(ordered):
        if i + 1 < len(ordered):
            info["end"] = ordered[i + 1][1]["start"] - 1
        else:
            info["end"] = ws.max_row

    return {
        "total_rows": total_rows,
        "sections": sections,
        "summary_rows": summary_rows,
        "geo_summary_rows": geo_summary_rows,
        "data_rows": data_rows,
    }


def _to_key(by_label: str) -> str:
    """Normalise 'By Product Type' -> 'by_product_type'."""
    cleaned = re.sub(r'^by\s+', '', by_label, flags=re.IGNORECASE).strip()
    return "by_" + re.sub(r'[^a-z0-9]+', '_', cleaned.lower()).strip('_')


def _classify_rows_in_section(section_info: dict, summary_rows: dict,
                              data_rows: list[int]) -> dict:
    """Classify data rows within a single section into subsections.

    Returns: {row: ("subsection_key", is_geographic)}
    """
    start = section_info["start"]
    end = section_info["end"]

    # Filter to rows in this section
    section_data_rows = [r for r in data_rows if start < r <= end]
    section_summaries = {r: label for r, label in summary_rows.items() if start < r <= end}

    if not section_data_rows:
        return {}

    # Find last summary row in this section
    summary_rows_sorted = sorted(section_summaries.keys())
    last_summary = max(summary_rows_sorted) if summary_rows_sorted else 0

    classifications = {}
    for row in section_data_rows:
        if row > last_summary:
            # After all "By ..." summaries → geographic
            classifications[row] = ("geographic", True)
        else:
            # Find nearest "By ..." summary below this row
            group_key = None
            for sr in summary_rows_sorted:
                if sr > row:
                    group_key = _to_key(section_summaries[sr])
                    break
            if group_key:
                classifications[row] = (group_key, False)
            else:
                classifications[row] = ("geographic", True)

    return classifications


# ─── Data Reading ────────────────────────────────────────────────────────────


def _round_val(val, decimals: int):
    """Round a value if it's a float."""
    if isinstance(val, float):
        return round(val, decimals)
    return val


def _read_year_group(ws, row: int, cols: list[tuple[int, int]], decimals: int = 2) -> dict:
    """Read a group of year columns for a row into {year_str: value}."""
    data = {}
    for col, year in cols:
        val = ws.cell(row=row, column=col).value
        if val is not None:
            data[str(year)] = _round_val(val, decimals)
    return data


def _read_row_full(ws, row: int, col_groups: dict, cagr_key: str,
                   include_pct: bool = True) -> dict:
    """Read all column groups for a single data row.

    Returns:
        {
            "forecast": {"2020": x, ..., "2032": x},
            "cagr_2025_2032": x,
            "yoy_growth": {"2021": x, ..., "2032": x},
            "percentage_share": {"2020": x, ..., "2032": x},  # only if include_pct and data exists
        }
    """
    result = {}

    # Group 1: Forecast
    forecast = _read_year_group(ws, row, col_groups["forecast_cols"], decimals=2)
    if forecast:
        result["forecast"] = forecast

    # CAGR
    if col_groups["cagr_col"]:
        cagr_val = ws.cell(row=row, column=col_groups["cagr_col"]).value
        if cagr_val is not None:
            result[cagr_key] = _round_val(cagr_val, 6)

    # Group 2: YoY Growth
    if col_groups["yoy_cols"]:
        yoy = _read_year_group(ws, row, col_groups["yoy_cols"], decimals=6)
        if yoy:
            result["yoy_growth"] = yoy

    # Group 3: Percentage Share
    if include_pct and col_groups["pct_cols"]:
        pct = _read_year_group(ws, row, col_groups["pct_cols"], decimals=6)
        if pct:
            result["percentage_share"] = pct

    return result


# ─── Sheet Extraction ────────────────────────────────────────────────────────


def _extract_sheet(ws, sheet_name: str) -> dict:
    """Extract all market data from one worksheet using discovery."""
    logger.info(f"Processing sheet: {sheet_name}")

    header_row = _discover_header_row(ws)
    if header_row is None:
        logger.warning(f"No header row found in {sheet_name}")
        return {}

    col_groups = _discover_column_groups(ws, header_row)

    if not col_groups["forecast_cols"]:
        logger.warning(f"No forecast columns in {sheet_name}")
        return {}

    cagr_key = _parse_cagr_key(col_groups["cagr_label"])
    years = [y for _, y in col_groups["forecast_cols"]]
    yoy_years = [y for _, y in col_groups["yoy_cols"]]
    pct_years = [y for _, y in col_groups["pct_cols"]]
    logger.info(
        f"  Forecast: {min(years)}-{max(years)} | "
        f"YoY: {min(yoy_years)}-{max(yoy_years) if yoy_years else 'none'} | "
        f"Pct: {min(pct_years)}-{max(pct_years) if pct_years else 'none'} | "
        f"CAGR: {cagr_key}"
    )

    structure = _discover_row_structure(ws)

    # Build row label map
    row_labels = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            row_labels[row] = str(val).strip()

    result = {}

    # ── Read totals (no percentage share) ─────────────────────────────────
    total = {}
    for kind in ("volume", "value"):
        tr = structure["total_rows"].get(kind)
        if tr:
            total[kind] = _read_row_full(ws, tr, col_groups, cagr_key, include_pct=False)
    if total:
        result["total"] = total

    # ── Read each section ─────────────────────────────────────────────────
    section_map = {
        "market_volume": "market_volume",
        "pricing": "pricing",
        "market_value": "market_value",
    }

    for section_key, output_key in section_map.items():
        section_info = structure["sections"].get(section_key)
        if not section_info:
            continue

        is_pricing = section_key == "pricing"

        classifications = _classify_rows_in_section(
            section_info,
            structure["summary_rows"],
            structure["data_rows"],
        )

        section_data = {"unit": section_info.get("unit", "")}

        for row, (subsection_key, is_geo) in sorted(classifications.items()):
            label = row_labels.get(row, "")
            if not label:
                continue

            # Pricing rows don't have percentage share
            include_pct = not is_pricing
            row_data = _read_row_full(ws, row, col_groups, cagr_key, include_pct=include_pct)
            if not row_data:
                continue

            # Geographic rows use "by_region" or "by_country" key
            if is_geo:
                subsection_key = "by_region"

            section_data.setdefault(subsection_key, {})[label] = row_data

        result[output_key] = section_data

    return result


# ─── Public API ──────────────────────────────────────────────────────────────


def extract_me(xlsx_path: str | Path) -> dict:
    """
    Extract Market Estimate data from an XLSX file.

    All structure is auto-discovered: header rows, year ranges, column groups
    (forecast, YoY growth, percentage share), segment dimensions, geographic
    entries, pricing sections, and CAGR periods.

    Args:
        xlsx_path: Path to the XLSX file.

    Returns:
        Dictionary with global and regional market data structured as:
        {
            "total": {volume/value with forecast + cagr + yoy},
            "market_volume": {by_product_type, by_config, ..., by_region — each with forecast + cagr + yoy + pct},
            "pricing": {by_product_type — each with forecast + cagr + yoy},
            "market_value": {by_product_type, by_config, ..., by_region — each with forecast + cagr + yoy + pct},
        }
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

    logger.info(f"Extracting ME data from: {xlsx_path.name}")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Step 1: Identify which sheets contain actual market data
    data_sheets = []
    skip_sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        if _is_data_sheet(ws):
            data_sheets.append(name)
        else:
            skip_sheets.append(name)

    logger.info(f"  Data sheets: {data_sheets}")
    logger.info(f"  Skipped sheets: {skip_sheets}")

    # Step 2: Identify the global sheet
    global_sheet = _identify_global_sheet(wb, data_sheets)
    logger.info(f"  Global sheet: {global_sheet}")

    # Step 2b: Determine which sheets are actual regional sheets
    valid_regional_sheets = set()
    if global_sheet:
        gs = wb[global_sheet]
        for row in range(1, gs.max_row + 1):
            val = gs.cell(row=row, column=1).value
            if val:
                label = str(val).strip()
                for ds in data_sheets:
                    if ds != global_sheet and ds.strip() == label:
                        valid_regional_sheets.add(ds)
    logger.info(f"  Valid regional sheets: {valid_regional_sheets}")

    # Step 3: Extract each sheet
    market_name = xlsx_path.stem
    for prefix in ("ME_", "ME ", "Market_Estimate_"):
        if market_name.startswith(prefix):
            market_name = market_name[len(prefix):]
    for suffix in ("_CMI", "_cmi"):
        if market_name.endswith(suffix):
            market_name = market_name[:-len(suffix)]
    market_name = market_name.replace("_", " ")

    sheets_to_process = [global_sheet] + [s for s in data_sheets if s in valid_regional_sheets]
    excluded_sheets = [s.strip() for s in data_sheets if s != global_sheet and s not in valid_regional_sheets]

    result = {
        "market_name": market_name,
        "source_file": xlsx_path.name,
        "sheets_found": wb.sheetnames,
        "data_sheets": [s.strip() for s in sheets_to_process],
        "skipped_sheets": [s.strip() for s in skip_sheets] + excluded_sheets,
        "global_sheet_name": global_sheet.strip() if global_sheet else None,
    }

    for name in sheets_to_process:
        ws = wb[name]
        clean = name.strip()
        sheet_data = _extract_sheet(ws, clean)
        if not sheet_data:
            continue

        if name == global_sheet:
            result["global"] = sheet_data
        else:
            result.setdefault("regions", {})[clean] = sheet_data

    wb.close()
    logger.info(f"Extracted {len(sheets_to_process)} sheets")
    return result
